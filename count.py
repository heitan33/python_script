# -*- coding: utf-8 -*-
import smtplib,commands,time,subprocess
import os
import threading
import openpyxl

wb = openpyxl.Workbook()
#wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
#for i in range(1,5):
#	wb.create_sheet()

def cpu_stat():
	wb.create_sheet(index=0,title='CPU-percent')
	sheet = wb.get_sheet_by_name('CPU-percent')
	wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
#	sheet['B1'] = 'CPU使用率'
#	sheet['A1'] = '日期'	
	while True:
		h=1
		i=1
		while True:
			date = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
			cpu = commands.getoutput("sar -u 1 1|sed -n '4p'|awk '{print $NF}'")
			cpu = 100 - float(cpu)
			time.sleep(3600)
			sheet['A' + str(h)] = date
			sheet['B' + str(h)] = cpu
			i=i+1
			h=h+1
			if (i>48):
				break
			else:
				j = time.strftime('%Y-%m-%d',time.localtime(time.time()))
				wb.save('%s.xlsx'%j)
				print j	


def free_stat():
	wb.create_sheet(index=1,title='memory-rest')
        sheet = wb.get_sheet_by_name('memory-rest')
#        sheet['B1'] = '内存余量'
#        sheet['A1'] = '日期'

	while True:
		h=1
		i=1
		while True:
			date = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
			buff=commands.getoutput("free -m|sed '1d'|grep buffers|awk '{print $4}'")	
			if buff == '':
				buff=commands.getoutput("free -m|sed -n '2p'|awk '{print $NF}'")
			time.sleep(3600)
                        sheet['A' + str(h)] = date
                        sheet['B' + str(h)] = buff
                        i=i+1
                        h=h+1
                        if (i>48):
                                break
                        else:
                                j = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                                wb.save('%s.xlsx'%j)
                                print j


def load_stat():
	wb.create_sheet(index=2,title='LOAD')
        sheet = wb.get_sheet_by_name('LOAD')
#        sheet['B1'] = '负载'
#        sheet['A1'] = '日期'

	while True:
		h=1
		i=1
		while True:
			try:
				date = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
				load=commands.getoutput("uptime|awk -F, '{print $(NF-1)}'")
				time.sleep(3600)
				sheet['A' + str(h)] = date
                	        sheet['B' + str(h)] = load
                	        i=i+1
                	        h=h+1
                	        if (i>48):
                	                break
                	        else:
                	                j = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                	                wb.save('%s.xlsx'%j)
                	                print j

			except:
				continue

def up_down():
	wb.create_sheet(index=3,title='tomcat\'s memory footprint')
        sheet = wb.get_sheet_by_name('tomcat\'s memory footprint')
#	sheet['B1'] = 'tomcat内存用量'
#	sheet['A1'] = '日期'

	while True:
		h=1
		i=1
		while True:
			try:
				date = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
				vin=commands.getoutput("pmap -d `netstat -ntulp|grep '8080'|awk '{print $7}'|awk -F/ '{print $1}'`|tail -n 1|awk '{print $4}'")
				time.sleep(3600)
				sheet['A' + str(h)] = date
				sheet['B' + str(h)] = vin
				i=i+1
				h=h+1
				if (i>48):
					break
				else:
					j = time.strftime('%Y-%m-%d',time.localtime(time.time()))
					wb.save('%s.xlsx'%j)
					print j
			except:
				continue

def links_count():
	wb.create_sheet(index=4,title='11573 LINKS')
	sheet = wb.get_sheet_by_name('11573 LINKS')
#        sheet['B1'] = '11573端口连接数'
#        sheet['A1'] = '日期'      
        while True:
                i=1
                h=1
                while True:
                        try:
                                date = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
                                vin=commands.getoutput("netstat -nat | grep -iw '11573' | wc -l")
                                time.sleep(3600)
                                sheet['A' + str(h)] = date
                                sheet['B' + str(h)] = vin
                                i=i+1
                                h=h+1
                                if (i>48):
                                        break
                                else:
                                        j = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                                        wb.save('%s.xlsx'%j)
                                        print j
                        except:
                                continue


def main():
	threading.Thread(target=cpu_stat).start()
	threading.Thread(target=free_stat).start()
	threading.Thread(target=load_stat).start()
	threading.Thread(target=up_down).start()
	threading.Thread(target=links_count).start()

if __name__== "__main__":
	main()
